import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions

def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}  # Dict: <class 'dict'>: {'firstname': 'Rahul', 'lastname': 'shetty'}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:\\Users\\vinhf\\Downloads\\download1.xlsx"
fruit_name = "Apple"
newValue = "999"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID, "downloadButton").click()


# call update_excel_data to edit the excel with update value
update_excel_data(file_path, fruit_name, "price", newValue)

#upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
assert driver.find_element(*toast_locator).text == "Updated Excel Data Successfully."

wait = WebDriverWait(driver,10)
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
wait = WebDriverWait(driver,10)
assert actual_price == newValue

time.sleep(5)
driver.close()

"""
#from selenium.webdriver.chrome import webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl
 
 
def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}
 
    for i in range(1,sheet.max_column+ 1):
        if sheet.cell(row=1,column=i).value == colName:
            Dict["col"] = i
 
    for i in range(1,sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i,column= j).value == searchTerm:
                Dict["row"] = i
 
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)
 
 
 
 
 
 
 
 
 
file_path = "/Users/rahulshetty/downloads/download.xlsx"
fruit_name = "Apple"
newValue = "990"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()
 
#edit the excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)
 
#upload
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)
 
wait = WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert actual_price == newValue
 
"""