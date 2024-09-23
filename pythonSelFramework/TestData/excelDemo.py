import openpyxl # import library to use excel file

book = openpyxl.load_workbook("C:\\Users\\vinhf\\PycharmProjects\\PythonDemo.xlsx")
sheet = book.active
Dict = {}  # Dict: <class 'dict'>: {'firstname': 'Rahul', 'lastname': 'shetty'}
cell = sheet.cell(row=1, column=2)
#print(cell.value)

# sheet.cell(row=2, column=2).value = "Rahul"
# print(sheet.cell(row=2, column=2).value)
# print(sheet.max_row)
#print(sheet.max_column)
#print(sheet['A5'].value)
#print(sheet['B3'].value)

for i in range(1, sheet.max_row+1):  # to get row
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):  # to get column
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
