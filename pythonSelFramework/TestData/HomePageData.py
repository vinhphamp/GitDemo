import openpyxl
class HomePageData:

    test_HomePage_data = [{"firstname": "Rahul", "lastname": "shetty", "gender": "Male"}, {"firstname": "Ashaki", "lastname": "shetty", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}  # Dict: <class 'dict'>: {'firstname': 'Rahul', 'lastname': 'shetty'}
        book = openpyxl.load_workbook("C:\\Users\\vinhf\\PycharmProjects\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # to get row
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get column
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
